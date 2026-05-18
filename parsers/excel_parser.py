"""Excel 解析器 — 基于真实文件布局的精确提取

支持 .xls (xlrd) 和 .xlsx (openpyxl) 两种格式。

解析策略: 6 阶段行列提取
  1. 元数据 (rows 0-6):  keyword:value 配对
  2. 仪器 (rows 7-8):    名称/编号/有效期/溯源方式
  3. 样品数据 (rows 11-18): 逐列映射 + 填报值解析
  4. FAS 标定 (rows 19-23): 日期/体积/浓度
  5. 页脚 (rows 24-27):   分析人/复核人/审核人
  6. Sheet 2 QC 数据:      全程序空白/实验室空白/平行样/标样
"""

import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from models.record import (
    CODRecord, SampleRow, Instrument, FASStandardization, QCData,
)
from parsers.field_extractor import (
    safe_float, safe_int, safe_date, clean_text,
    parse_reported_cod, classify_samples,
    parse_guarantee_range,
)
from parsers.excel_layout import (
    SHEET1_KEYWORDS, SHEET2_KEYWORDS,
    METADATA_KEYWORDS, SAMPLE_COL_MAP,
    INSTRUMENT_ROW_START, INSTRUMENT_ROW_END,
    FAS_DATE_ROW, FAS_VOL_ROW2, FAS_VOL_ROW3,
    FOOTER_KEYWORDS,
    QC_FIELD_BLANK_ROW, QC_FIELD_BLANK_COL_ID,
    QC_FIELD_BLANK_COL_GUARANTEE, QC_FIELD_BLANK_COL_MEASURED,
    QC_FIELD_BLANK_COL_QUALIFIED,
    QC_LAB_BLANK_START, QC_LAB_BLANK_END,
    QC_PARALLEL_ROW,
    QC_PARALLEL_COL_ID, QC_PARALLEL_COL_VAL1, QC_PARALLEL_COL_VAL2,
    QC_PARALLEL_COL_MEAN, QC_PARALLEL_COL_RPD, QC_PARALLEL_COL_QUALIFIED,
    QC_STANDARD_ROW,
    QC_STANDARD_COL_ID, QC_STANDARD_COL_GUARANTEE,
    QC_STANDARD_COL_MEASURED, QC_STANDARD_COL_QUALIFIED,
)


# ============================================================
# 公共入口
# ============================================================

def parse_excel(file_path: str) -> CODRecord:
    """解析 Excel 文件为 CODRecord (保持向后兼容)"""
    ext = Path(file_path).suffix.lower()
    if ext == '.xls':
        return _parse_xls(file_path)
    elif ext in ('.xlsx', '.xlsm'):
        return _parse_xlsx(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def parse_excel_bytes(file_bytes: bytes, filename: str) -> CODRecord:
    """从字节流解析 Excel (供 Streamlit 上传使用)"""
    ext = Path(filename).suffix.lower()
    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        return _extract_from_xlrd(wb)
    elif ext in ('.xlsx', '.xlsm'):
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        rec = _extract_from_openpyxl(wb)
        wb.close()
        return rec
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


# ============================================================
# .xls 支持
# ============================================================

def _parse_xls(file_path: str) -> CODRecord:
    import xlrd
    wb = xlrd.open_workbook(file_path)
    return _extract_from_xlrd(wb)


def _extract_from_xlrd(wb) -> CODRecord:
    """从 xlrd Workbook 完整提取"""
    rec = CODRecord()

    # 找 Sheet 1 和 Sheet 2
    sheet1 = _find_sheet_xlrd(wb, SHEET1_KEYWORDS)
    sheet2 = _find_sheet_xlrd(wb, SHEET2_KEYWORDS)

    if sheet1 is None:
        sheet1 = wb.sheet_by_index(0)

    _extract_metadata_xlrd(sheet1, rec)
    _extract_instruments_xlrd(sheet1, rec)
    _extract_samples_xlrd(sheet1, rec)
    _extract_fas_xlrd(sheet1, rec)
    _extract_footer_xlrd(sheet1, rec)

    # Sheet 2 QC 数据
    if sheet2 is not None:
        _extract_qc_sheet_xlrd(sheet2, rec)
    elif wb.nsheets > 1:
        sheet2 = wb.sheet_by_index(1)
        _extract_qc_sheet_xlrd(sheet2, rec)

    # 后处理: 样品分类
    classify_samples(rec.samples)

    # 自动检测方法级别
    _detect_method_level(rec)

    return rec


def _find_sheet_xlrd(wb, keywords: list[str]):
    """按关键词查找 Sheet"""
    for k in keywords:
        for i in range(wb.nsheets):
            name = wb.sheet_names()[i]
            if k in name:
                return wb.sheet_by_index(i)
    return None


# ============================================================
# .xlsx 支持
# ============================================================

def _parse_xlsx(file_path: str) -> CODRecord:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    rec = _extract_from_openpyxl(wb)
    wb.close()
    return rec


def _extract_from_openpyxl(wb) -> CODRecord:
    """从 openpyxl Workbook 完整提取"""
    rec = CODRecord()

    sheet1 = _find_sheet_openpyxl(wb, SHEET1_KEYWORDS)
    sheet2 = _find_sheet_openpyxl(wb, SHEET2_KEYWORDS)

    if sheet1 is None:
        sheet1 = wb.worksheets[0]

    _extract_metadata_openpyxl(sheet1, rec)
    _extract_instruments_openpyxl(sheet1, rec)
    _extract_samples_openpyxl(sheet1, rec)
    _extract_fas_openpyxl(sheet1, rec)
    _extract_footer_openpyxl(sheet1, rec)

    if sheet2 is not None:
        _extract_qc_sheet_openpyxl(sheet2, rec)
    elif len(wb.worksheets) > 1:
        _extract_qc_sheet_openpyxl(wb.worksheets[1], rec)

    classify_samples(rec.samples)
    _detect_method_level(rec)

    return rec


def _find_sheet_openpyxl(wb, keywords: list[str]):
    for k in keywords:
        for ws in wb.worksheets:
            if k in (ws.title or ''):
                return ws
    return None


# ============================================================
# 阶段 1: 元数据提取
# ============================================================

def _extract_metadata_xlrd(sheet, rec: CODRecord):
    """扫描前 7 行提取元数据"""
    for row_idx in range(min(sheet.nrows, 7)):
        for col_idx in range(sheet.ncols):
            val = str(sheet.cell_value(row_idx, col_idx)).strip()
            for keyword, attr in METADATA_KEYWORDS.items():
                if keyword in val:
                    # 取同行下一列或隔列的值 (org_name/task_name 可能较远, 最高 offset 6)
                    for offset in (1, 2, 3, 4, 5, 6):
                        if col_idx + offset < sheet.ncols:
                            next_val = str(sheet.cell_value(row_idx, col_idx + offset)).strip()
                            if next_val and keyword.strip() not in next_val:
                                _set_metadata_field(rec, attr, next_val)
                                break


def _extract_metadata_openpyxl(sheet, rec: CODRecord):
    """openpyxl 版本"""
    for row in sheet.iter_rows(min_row=1, max_row=7):
        cells = [(c.column - 1, str(c.value).strip() if c.value is not None else '')
                 for c in row]
        for col_idx, val in cells:
            for keyword, attr in METADATA_KEYWORDS.items():
                if keyword in val:
                    for offset in (1, 2, 3):
                        for oc, ov in cells:
                            if oc == col_idx + offset and ov and keyword.strip() not in ov:
                                _set_metadata_field(rec, attr, ov)
                                break


def _set_metadata_field(rec: CODRecord, attr: str, val: str):
    """设置元数据字段"""
    if attr == 'record_id':
        rec.record_id = val
    elif attr == 'task_id':
        rec.task_id = val
    elif attr == 'org_name':
        rec.org_name = val or rec.org_name
    elif attr == 'task_name':
        rec.task_name = val or rec.task_name
    elif attr == 'sampling_date':
        d = safe_date(val)
        if d:
            rec.sampling_date = d
    elif attr == 'analysis_date':
        d = safe_date(val)
        if d:
            rec.analysis_date = d
    elif attr == 'method_ref':
        rec.method_ref = val
    elif attr == 'temperature':
        rec.temperature = safe_float(val)
    elif attr == 'humidity':
        rec.humidity = safe_float(val)
    elif attr == 'k2cr2o7_prep_date':
        d = safe_date(val)
        if d:
            rec.k2cr2o7_prep_date = d
    elif attr == 'k2cr2o7_conc':
        rec.k2cr2o7_conc = safe_float(val)


# ============================================================
# 阶段 2: 仪器提取
# ============================================================

def _extract_instruments_xlrd(sheet, rec: CODRecord):
    """从 rows 7-8 提取仪器"""
    for row_idx in range(INSTRUMENT_ROW_START, min(sheet.nrows, INSTRUMENT_ROW_END + 1)):
        cells = {}
        for col_idx in range(sheet.ncols):
            v = str(sheet.cell_value(row_idx, col_idx)).strip()
            if v:
                cells[col_idx] = v

        if not cells:
            continue

        inst = Instrument()

        # 列 0-1: 名称/型号
        for c in (0, 1):
            if c in cells and cells[c] not in ('仪器及型号', '仪器编号', ''):
                name = cells[c]
                if inst.name:
                    inst.model = name
                else:
                    inst.name = name

        # 列 2: "仪器编号" 标签
        # 列 3-4: 编号值
        for c in (3, 4):
            if c in cells and cells[c] not in ('仪器编号',):
                inst.serial_no = cells[c]
                break

        # 列 5: "仪器溯源有效期" 标签
        # 列 6-7: 日期值
        for c in (6, 7):
            if c in cells:
                d = safe_date(cells[c])
                if d:
                    inst.calibration_expiry = d
                    break

        # 列 8: "仪器溯源方式" 标签
        # 列 9-11: 溯源方式
        for c in (9, 10, 11):
            if c in cells and '仪器' not in cells[c] and '溯源' not in cells[c]:
                inst.calibration_method = cells[c]
                break

        if inst.name:
            rec.instruments.append(inst)


def _extract_instruments_openpyxl(sheet, rec: CODRecord):
    """openpyxl 版本 — 同样逻辑"""
    row_indices = list(range(INSTRUMENT_ROW_START, INSTRUMENT_ROW_END + 1))
    for row_idx, row in enumerate(sheet.iter_rows(min_row=INSTRUMENT_ROW_START + 1,
                                                   max_row=INSTRUMENT_ROW_END + 1)):
        cells = {}
        for cell in row:
            v = str(cell.value).strip() if cell.value is not None else ''
            if v:
                cells[cell.column - 1] = v

        if not cells:
            continue

        inst = Instrument()
        for c in (0, 1):
            if c in cells and cells[c] not in ('仪器及型号', '仪器编号', ''):
                if inst.name:
                    inst.model = cells[c]
                else:
                    inst.name = cells[c]

        for c in (3, 4):
            if c in cells and cells[c] not in ('仪器编号',):
                inst.serial_no = cells[c]
                break

        for c in (6, 7):
            if c in cells:
                d = safe_date(cells[c])
                if d:
                    inst.calibration_expiry = d
                    break

        for c in (9, 10, 11):
            if c in cells and '仪器' not in cells[c] and '溯源' not in cells[c]:
                inst.calibration_method = cells[c]
                break

        if inst.name:
            rec.instruments.append(inst)


# ============================================================
# 阶段 3: 样品数据提取
# ============================================================

def _extract_samples_xlrd(sheet, rec: CODRecord):
    """从 rows 11+ 提取样品数据"""
    # 动态检测样品数据起始行 (查找包含 "1" 作为序号的第一个数据行)
    data_start = None
    for row_idx in range(9, sheet.nrows):
        seq_val = str(sheet.cell_value(row_idx, 0)).strip()
        if seq_val == '1' and row_idx > 9:
            data_start = row_idx
            break

    if data_start is None:
        return

    # 读取所有数据行，直到遇到空行/标定区/页脚
    for row_idx in range(data_start, sheet.nrows):
        seq_val = str(sheet.cell_value(row_idx, 0)).strip()

        # 停止条件: 空序号、标定区、说明、备注、分析人
        if not seq_val or seq_val in ('/',):
            break
        if '硫酸亚铁铵' in seq_val:
            break
        if seq_val in ('说明', '备注', '分析人', '复核人', '审核人'):
            break

        # 非数字序号 → 可能是没有填写的空白行，跳过
        try:
            int(seq_val)
        except ValueError:
            continue

        sample = SampleRow(seq=int(seq_val))

        # 按 SAMPLE_COL_MAP 映射
        for col_idx, field_name in SAMPLE_COL_MAP.items():
            if col_idx < sheet.ncols:
                raw_val = str(sheet.cell_value(row_idx, col_idx)).strip()
                if raw_val:
                    _set_sample_field(sample, field_name, raw_val)

        # 如果 sample_id 为空，跳过
        if not sample.sample_id:
            continue

        rec.samples.append(sample)

    # 找 HgSO4 加入量 (可能在 col 12+)
    # 真实文件中 HgSO4 在单独区域，不在主表
    _extract_hgso4_xlrd(sheet, rec)


def _extract_hgso4_xlrd(sheet, rec):
    """从表格中提取 HgSO4 加入量"""
    # 在真实文件中，硫酸汞加入量可能在第 12 列之后
    # 从行 11 开始扫描
    for row_idx in range(11, sheet.nrows):
        if row_idx < len(rec.samples):
            sample_idx = row_idx - 11  # 相对于数据起始行
            # 检查是否有额外列
            for col_idx in range(12, min(sheet.ncols, 15)):
                cell = str(sheet.cell_value(row_idx, col_idx)).strip()
                if cell and cell not in ('/', '-', ''):
                    val = safe_float(cell)
                    if val is not None and 0 < sample_idx < len(rec.samples) + 1:
                        # 尝试映射: 可能 HgSO4 在 col 12
                        if col_idx == 12 and 0 <= sample_idx < len(rec.samples):
                            rec.samples[sample_idx].hgso4_added = val


def _extract_samples_openpyxl(sheet, rec: CODRecord):
    """openpyxl 版本"""
    rows = list(sheet.iter_rows())
    data_start = None

    for row_idx, row in enumerate(rows):
        if row_idx < 9:
            continue
        seq_val = str(row[0].value).strip() if row[0].value is not None else ''
        if seq_val == '1' and row_idx > 9:
            data_start = row_idx
            break

    if data_start is None:
        return

    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        seq_val = str(row[0].value).strip() if row[0].value is not None else ''
        if not seq_val or seq_val in ('/'):
            break
        if '硫酸亚铁铵' in seq_val:
            break
        if seq_val in ('说明', '备注', '分析人', '复核人', '审核人'):
            break

        try:
            int(seq_val)
        except ValueError:
            continue

        sample = SampleRow(seq=int(seq_val))

        for col_idx, field_name in SAMPLE_COL_MAP.items():
            if col_idx < len(row):
                cell = row[col_idx]
                raw_val = str(cell.value).strip() if cell.value is not None else ''
                if raw_val:
                    _set_sample_field(sample, field_name, raw_val)

        if not sample.sample_id:
            continue

        rec.samples.append(sample)


def _set_sample_field(sample: SampleRow, field_name: str, raw: str):
    """设置单个样品字段"""
    raw = clean_text(raw)

    if field_name == 'sample_id':
        sample.sample_id = raw
    elif field_name == 'volume':
        sample.volume = safe_float(raw) or 10.00
    elif field_name == 'dilution_factor':
        sample.dilution_factor = safe_float(raw) or 1.0
    elif field_name == 'diluted_volume':
        sample.diluted_volume = safe_float(raw) or 10.00
    elif field_name == 'k2cr2o7_conc':
        sample.k2cr2o7_conc = safe_float(raw) or 0.0250
    elif field_name == 'end_reading':
        sample.end_reading = safe_float(raw) or 0.0
    elif field_name == 'start_reading':
        sample.start_reading = safe_float(raw) or 0.0
    elif field_name == 'net_volume':
        sample.net_volume = safe_float(raw) or 0.0
    elif field_name == 'reported_cod_raw':
        sample.reported_cod_raw = raw
        cod_val, is_below = parse_reported_cod(raw)
        sample.reported_cod = cod_val
        sample.is_below_dl = is_below
    elif field_name == 'salinity':
        sample.salinity = safe_float(raw)
    elif field_name == 'cl_estimate':
        sample.cl_estimate = safe_float(raw)


# ============================================================
# 阶段 4: FAS 标定
# ============================================================

def _extract_fas_xlrd(sheet, rec: CODRecord):
    """从 rows 19-23 提取 FAS 标定数据"""
    fas = rec.fas_std

    # 标定日期 + 重铬酸钾用量 + 浓度 — row 21
    if FAS_DATE_ROW < sheet.nrows:
        for col_idx in range(sheet.ncols):
            val = str(sheet.cell_value(FAS_DATE_ROW, col_idx)).strip()
            if col_idx == 1 and val:
                d = safe_date(val)
                if d:
                    fas.date = d
            elif col_idx in (2, 3) and val and safe_float(val) is not None:
                fas.k2cr2o7_volume = safe_float(val) or 5.00
            elif col_idx in (4, 5) and val and safe_float(val) is not None:
                fas.k2cr2o7_conc = safe_float(val) or 0.0250

    # 平行滴定体积 — row 21, 22, 23 (col 6-8)
    for row_idx in (FAS_DATE_ROW, FAS_VOL_ROW2, FAS_VOL_ROW3):
        if row_idx >= sheet.nrows:
            continue
        # 优先取净用量列 (col 8)
        for col_idx in (8, 6, 7):
            if col_idx < sheet.ncols:
                val = str(sheet.cell_value(row_idx, col_idx)).strip()
                f = safe_float(val)
                if f is not None and f > 0:
                    if col_idx == 8:  # 净用量 — 这是实际标定体积
                        fas.volumes.append(f)
                    elif not fas.volumes:  # 仅在净用量为空时用终读数
                        fas.volumes.append(f)
                    break

    # 标定浓度 — row 21 col 9
    if FAS_DATE_ROW < sheet.nrows and 9 < sheet.ncols:
        conc_val = str(sheet.cell_value(FAS_DATE_ROW, 9)).strip()
        fas.reported_conc = safe_float(conc_val)


def _extract_fas_openpyxl(sheet, rec: CODRecord):
    """openpyxl 版本"""
    fas = rec.fas_std
    rows = list(sheet.iter_rows())

    if FAS_DATE_ROW < len(rows):
        row = rows[FAS_DATE_ROW]
        for cell in row:
            col = cell.column - 1
            val = str(cell.value).strip() if cell.value is not None else ''
            if col == 1 and val:
                d = safe_date(val)
                if d:
                    fas.date = d
            elif col in (2, 3) and val and safe_float(val) is not None:
                fas.k2cr2o7_volume = safe_float(val) or 5.00
            elif col in (4, 5) and val and safe_float(val) is not None:
                fas.k2cr2o7_conc = safe_float(val) or 0.0250

    for row_idx in (FAS_DATE_ROW, FAS_VOL_ROW2, FAS_VOL_ROW3):
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        for cell in row:
            col = cell.column - 1
            if col in (8, 6, 7):
                val = str(cell.value).strip() if cell.value is not None else ''
                f = safe_float(val)
                if f is not None and f > 0:
                    if col == 8:
                        fas.volumes.append(f)
                    elif not fas.volumes:
                        fas.volumes.append(f)
                    break

    if FAS_DATE_ROW < len(rows) and 9 < len(rows[FAS_DATE_ROW]):
        val = str(rows[FAS_DATE_ROW][9].value).strip()
        fas.reported_conc = safe_float(val)


# ============================================================
# 阶段 5: 页脚 (分析人/复核人/审核人)
# ============================================================

def _extract_footer_xlrd(sheet, rec: CODRecord):
    """从 rows 24+ 提取签字信息"""
    for row_idx in range(24, sheet.nrows):
        for col_idx in range(sheet.ncols):
            val = str(sheet.cell_value(row_idx, col_idx)).strip()
            if not val:
                continue
            for keyword, attr in FOOTER_KEYWORDS.items():
                if keyword in val:
                    # 取值: "分析人：苏毅" → "苏毅"
                    parts = val.split('：') if '：' in val else val.split(':')
                    if len(parts) > 1 and parts[1].strip():
                        name = parts[1].strip()
                    elif col_idx + 1 < sheet.ncols:
                        name = str(sheet.cell_value(row_idx, col_idx + 1)).strip()
                    else:
                        name = ''

                    if attr == 'analyst':
                        rec.analyst = name
                    elif attr == 'reviewer':
                        rec.reviewer = name
                    elif attr == 'approver':
                        rec.approver = name


def _extract_footer_openpyxl(sheet, rec: CODRecord):
    """openpyxl 版本"""
    rows = list(sheet.iter_rows())
    for row_idx in range(24, min(len(rows), 30)):
        for cell in rows[row_idx]:
            val = str(cell.value).strip() if cell.value is not None else ''
            if not val:
                continue
            for keyword, attr in FOOTER_KEYWORDS.items():
                if keyword in val:
                    parts = val.split('：') if '：' in val else val.split(':')
                    if len(parts) > 1 and parts[1].strip():
                        name = parts[1].strip()
                    elif cell.column < len(rows[row_idx]):
                        next_cell = rows[row_idx][cell.column]
                        name = str(next_cell.value).strip() if next_cell.value else ''
                    else:
                        name = ''
                    setattr(rec, attr, name)


# ============================================================
# 阶段 6: Sheet 2 QC 数据
# ============================================================

def _extract_qc_sheet_xlrd(sheet, rec: CODRecord):
    """从 Sheet 2 提取质控数据"""
    qc = rec.qc

    # 全程序空白 (row 4)
    if QC_FIELD_BLANK_ROW < sheet.nrows:
        row = QC_FIELD_BLANK_ROW
        qc.field_blank_sample_id = _cell_xlrd(sheet, row, QC_FIELD_BLANK_COL_ID)
        qc.field_blank_guarantee = _cell_xlrd(sheet, row, QC_FIELD_BLANK_COL_GUARANTEE)
        qc.field_blank_measured = _cell_xlrd(sheet, row, QC_FIELD_BLANK_COL_MEASURED)
        qc.field_blank_qualified = _is_qualified(_cell_xlrd(sheet, row, QC_FIELD_BLANK_COL_QUALIFIED))

    # 实验室空白列表 (rows 6-7)
    qc.lab_blank_sample_ids = []
    for row_idx in range(QC_LAB_BLANK_START, QC_LAB_BLANK_END + 1):
        if row_idx < sheet.nrows:
            sid = _cell_xlrd(sheet, row_idx, 0)
            if sid and sid != '/':
                qc.lab_blank_sample_ids.append(sid)

    # 平行样 (row 9)
    if QC_PARALLEL_ROW < sheet.nrows:
        row = QC_PARALLEL_ROW
        qc.parallel_sample_id = _cell_xlrd(sheet, row, QC_PARALLEL_COL_ID)
        qc.parallel_value1 = safe_float(_cell_xlrd(sheet, row, QC_PARALLEL_COL_VAL1))
        qc.parallel_value2 = safe_float(_cell_xlrd(sheet, row, QC_PARALLEL_COL_VAL2))
        qc.parallel_mean = safe_float(_cell_xlrd(sheet, row, QC_PARALLEL_COL_MEAN))
        qc.parallel_rpd = safe_float(_cell_xlrd(sheet, row, QC_PARALLEL_COL_RPD))
        qc.parallel_qualified = _is_qualified(_cell_xlrd(sheet, row, QC_PARALLEL_COL_QUALIFIED))

    # 有证标样 (row 11)
    if QC_STANDARD_ROW < sheet.nrows:
        row = QC_STANDARD_ROW
        qc.std_sample_id = _cell_xlrd(sheet, row, QC_STANDARD_COL_ID)
        qc.std_guarantee_range = _cell_xlrd(sheet, row, QC_STANDARD_COL_GUARANTEE)
        # 测量值可能是 "14.4(mg/L)" 格式
        measured_raw = _cell_xlrd(sheet, row, QC_STANDARD_COL_MEASURED)
        qc.std_measured = safe_float(measured_raw) or safe_float(measured_raw.split('(')[0] if '(' in measured_raw else measured_raw)
        qc.std_qualified = _is_qualified(_cell_xlrd(sheet, row, QC_STANDARD_COL_QUALIFIED))

    # 加标回收 (row 13, 通常为空) — 跳过


def _extract_qc_sheet_openpyxl(sheet, rec: CODRecord):
    """openpyxl 版本"""
    qc = rec.qc
    rows = list(sheet.iter_rows())

    def cell_val(r, c, default=''):
        if r < len(rows) and c < len(rows[r]):
            return str(rows[r][c].value).strip() if rows[r][c].value is not None else default
        return default

    qc.field_blank_sample_id = cell_val(QC_FIELD_BLANK_ROW, QC_FIELD_BLANK_COL_ID)
    qc.field_blank_guarantee = cell_val(QC_FIELD_BLANK_ROW, QC_FIELD_BLANK_COL_GUARANTEE)
    qc.field_blank_measured = cell_val(QC_FIELD_BLANK_ROW, QC_FIELD_BLANK_COL_MEASURED)
    qc.field_blank_qualified = _is_qualified(cell_val(QC_FIELD_BLANK_ROW, QC_FIELD_BLANK_COL_QUALIFIED))

    qc.lab_blank_sample_ids = []
    for row_idx in range(QC_LAB_BLANK_START, QC_LAB_BLANK_END + 1):
        sid = cell_val(row_idx, 0)
        if sid and sid != '/':
            qc.lab_blank_sample_ids.append(sid)

    qc.parallel_sample_id = cell_val(QC_PARALLEL_ROW, QC_PARALLEL_COL_ID)
    qc.parallel_value1 = safe_float(cell_val(QC_PARALLEL_ROW, QC_PARALLEL_COL_VAL1))
    qc.parallel_value2 = safe_float(cell_val(QC_PARALLEL_ROW, QC_PARALLEL_COL_VAL2))
    qc.parallel_mean = safe_float(cell_val(QC_PARALLEL_ROW, QC_PARALLEL_COL_MEAN))
    qc.parallel_rpd = safe_float(cell_val(QC_PARALLEL_ROW, QC_PARALLEL_COL_RPD))
    qc.parallel_qualified = _is_qualified(cell_val(QC_PARALLEL_ROW, QC_PARALLEL_COL_QUALIFIED))

    qc.std_sample_id = cell_val(QC_STANDARD_ROW, QC_STANDARD_COL_ID)
    qc.std_guarantee_range = cell_val(QC_STANDARD_ROW, QC_STANDARD_COL_GUARANTEE)
    measured_raw = cell_val(QC_STANDARD_ROW, QC_STANDARD_COL_MEASURED)
    qc.std_measured = safe_float(measured_raw) or safe_float(measured_raw.split('(')[0] if '(' in measured_raw else measured_raw)
    qc.std_qualified = _is_qualified(cell_val(QC_STANDARD_ROW, QC_STANDARD_COL_QUALIFIED))


# ============================================================
# 辅助函数
# ============================================================

def _cell_xlrd(sheet, row: int, col: int) -> str:
    """安全读取 xlrd 单元格"""
    if row < sheet.nrows and col < sheet.ncols:
        return clean_text(str(sheet.cell_value(row, col)))
    return ""


def _is_qualified(val: str) -> bool:
    """判断 '合格' → True, '不合格' → False, 其他 → True"""
    if not val:
        return True
    if '不合格' in val:
        return False
    return True


def _detect_method_level(rec: CODRecord):
    """自动检测方法级别"""
    conc = rec.k2cr2o7_conc
    if conc is None:
        # 从样品推断
        for s in rec.samples:
            if s.k2cr2o7_conc and s.k2cr2o7_conc > 0:
                conc = s.k2cr2o7_conc
                break

    if conc is not None:
        if abs(conc - 0.0250) < 0.001:
            for s in rec.samples:
                s.method_level = "low"
        elif abs(conc - 0.250) < 0.001:
            for s in rec.samples:
                s.method_level = "high"


# ============================================================
# 保留 build_demo_record() — 用于 UI 演示和数据对比
# ============================================================

def build_demo_record() -> CODRecord:
    """构建示例 CODRecord（真实样本数据）"""
    rec = CODRecord(
        record_id="GJW-04-2016-YS-SZ-011",
        task_id="202604450900",
        org_name="广西壮族自治区玉林生态环境监测中心",
        task_name="202604分析任务",
        sampling_date=date(2026, 4, 2),
        analysis_date=date(2026, 4, 3),
        method_ref="水质 化学需氧量的测定 重铬酸盐法(HJ 828-2017)",
        temperature=24.5,
        humidity=55.0,
        k2cr2o7_prep_date=date(2026, 4, 3),
        k2cr2o7_conc=0.0250,
        analyst="苏毅",
        reviewer="陈昶",
        approver="吴雄平",
    )

    rec.instruments = [
        Instrument(name="酸式滴定管(50ml)", model="", serial_no="YL-D50-001",
                   calibration_expiry=date(2026, 7, 4), calibration_method="检定"),
        Instrument(name="CODcr回流消解仪(1200K型)", model="1200K型", serial_no="07905",
                   calibration_expiry=date(2026, 5, 19), calibration_method=""),
    ]

    rec.fas_std = FASStandardization(
        date=date(2026, 4, 3),
        k2cr2o7_volume=5.00,
        k2cr2o7_conc=0.0250,
        volumes=[24.00, 24.04],
        reported_conc=0.005204,
    )

    rec.samples = [
        SampleRow(seq=1, sample_id="实验室空白-202604-000012",
                  volume=10.00, dilution_factor=1.0, diluted_volume=10.00,
                  k2cr2o7_conc=0.0250,
                  end_reading=22.60, start_reading=0.00, net_volume=22.60,
                  reported_cod_raw="/", reported_cod=None,
                  is_blank=True, cl_estimate=0.0, method_level="low"),
        SampleRow(seq=2, sample_id="实验室空白-202604-000013",
                  volume=10.00, dilution_factor=1.0, diluted_volume=10.00,
                  k2cr2o7_conc=0.0250,
                  end_reading=22.65, start_reading=0.00, net_volume=22.65,
                  reported_cod_raw="/", reported_cod=None,
                  is_blank=True, cl_estimate=0.0, method_level="low"),
        SampleRow(seq=3, sample_id="实验室标样-YLB20260289",
                  volume=10.00, dilution_factor=1.0, diluted_volume=10.00,
                  k2cr2o7_conc=0.0250,
                  end_reading=19.16, start_reading=0.00, net_volume=19.16,
                  reported_cod_raw="14", reported_cod=14.0,
                  is_qc_standard=True, cl_estimate=0.0, method_level="low"),
        SampleRow(seq=4, sample_id="12026040115513002029",
                  volume=10.00, dilution_factor=1.0, diluted_volume=10.00,
                  k2cr2o7_conc=0.0250,
                  end_reading=19.02, start_reading=0.00, net_volume=19.02,
                  reported_cod_raw="15", reported_cod=15.0,
                  cl_estimate=20.0, hgso4_added=0.04, method_level="low"),
        SampleRow(seq=5, sample_id="12026040115513002029(空白1)",
                  volume=10.00, dilution_factor=1.0, diluted_volume=10.00,
                  k2cr2o7_conc=0.0250,
                  end_reading=22.45, start_reading=0.00, net_volume=22.45,
                  reported_cod_raw="4L", reported_cod=None, is_below_dl=True,
                  cl_estimate=0.0, method_level="low"),
        SampleRow(seq=6, sample_id="12026040115510008172",
                  volume=10.00, dilution_factor=1.0, diluted_volume=10.00,
                  k2cr2o7_conc=0.0250,
                  end_reading=18.88, start_reading=0.00, net_volume=18.88,
                  reported_cod_raw="16", reported_cod=16.0,
                  cl_estimate=20.0, hgso4_added=0.04, method_level="low"),
        SampleRow(seq=7, sample_id="12026040115510012752",
                  volume=10.00, dilution_factor=1.0, diluted_volume=10.00,
                  k2cr2o7_conc=0.0250,
                  end_reading=18.74, start_reading=0.00, net_volume=18.74,
                  reported_cod_raw="16", reported_cod=16.0,
                  is_parallel=True, parallel_pair_id="12026040115510012752",
                  cl_estimate=20.0, hgso4_added=0.04, method_level="low"),
        SampleRow(seq=8, sample_id="12026040115510012752-1-平行",
                  volume=10.00, dilution_factor=1.0, diluted_volume=10.00,
                  k2cr2o7_conc=0.0250,
                  end_reading=18.70, start_reading=0.00, net_volume=18.70,
                  reported_cod_raw="16", reported_cod=16.0,
                  is_parallel=True, parallel_pair_id="12026040115510012752",
                  cl_estimate=20.0, hgso4_added=0.04, method_level="low"),
    ]

    rec.qc = QCData(
        field_blank_sample_id="12026040115513002029(空白1)",
        field_blank_guarantee="<4(mg/L)",
        field_blank_measured="4L",
        field_blank_qualified=True,
        lab_blank_sample_ids=["实验室空白-202604-000012", "实验室空白-202604-000013"],
        parallel_sample_id="12026040115510012752",
        parallel_value1=16.0,
        parallel_value2=16.0,
        parallel_mean=16.0,
        parallel_rpd=0.0,
        parallel_qualified=True,
        std_sample_id="实验室标样-YLB20260289",
        std_guarantee_range="14.3 ± 1.1(mg/L)",
        std_measured=14.4,
        std_qualified=True,
    )

    return rec
